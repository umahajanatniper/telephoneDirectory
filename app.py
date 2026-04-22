from __future__ import annotations

import os
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
from typing import Any

from flask import Flask, jsonify, render_template, request
from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
WORKBOOK_GLOB = "*.xlsx"
MAX_RESULTS = 50
HEADER_SCAN_LIMIT = 10


@dataclass(frozen=True)
class DirectoryMetadata:
    workbook_name: str
    sheet_name: str
    total_records: int
    display_columns: list[str]


def create_app() -> Flask:
    app = Flask(__name__)

    @app.get("/")
    def index() -> str:
        records, metadata = load_directory()
        return render_template(
            "index.html",
            metadata=metadata,
            initial_results=records[:12],
        )

    @app.get("/api/search")
    def search() -> Any:
        query = request.args.get("q", "").strip()
        records, metadata = load_directory()
        matches = search_records(records, query)
        return jsonify(
            {
                "query": query,
                "count": len(matches),
                "results": matches[:MAX_RESULTS],
                "metadata": {
                    "workbook": metadata.workbook_name,
                    "sheet": metadata.sheet_name,
                    "total_records": metadata.total_records,
                    "display_columns": metadata.display_columns,
                },
            }
        )

    return app


def find_workbook() -> Path:
    candidates = sorted(
        path
        for path in BASE_DIR.glob(WORKBOOK_GLOB)
        if not path.name.startswith("~$") and path.is_file()
    )
    if not candidates:
        raise FileNotFoundError("No Excel workbook was found in the project directory.")
    return candidates[0]


def normalize_header(value: Any) -> str:
    text = stringify(value).strip()
    return " ".join(text.lower().replace("_", " ").split())


def stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def choose_header_row(rows: list[tuple[Any, ...]]) -> tuple[int, list[str]]:
    for index, row in enumerate(rows[:HEADER_SCAN_LIMIT]):
        cleaned = [stringify(cell) for cell in row]
        populated = [cell for cell in cleaned if cell]
        if len(populated) >= 2:
            headers = [cell if cell else f"Column {position + 1}" for position, cell in enumerate(cleaned)]
            return index, headers
    raise ValueError("Unable to detect a usable header row in the workbook.")


def infer_role_column(headers: list[str]) -> str | None:
    """Return the first column whose header looks like a role/department field."""
    role_keywords = ("role", "department", "dept", "designation", "section", "office", "division")
    for header in headers:
        norm = normalize_header(header)
        if any(kw in norm for kw in role_keywords):
            return header
    return None


def infer_name_columns(headers: list[str]) -> list[str]:
    prioritized_keywords = (
        "name",
        "employee",
        "faculty",
        "person",
        "official",
        "contact",
    )
    normalized = {header: normalize_header(header) for header in headers}
    explicit = [
        header
        for header in headers
        if any(keyword in normalized[header] for keyword in prioritized_keywords)
    ]
    if explicit:
        return explicit

    fallback = [
        header
        for header in headers
        if not any(keyword in normalized[header] for keyword in ("phone", "mobile", "email", "ext", "landline"))
    ]
    return fallback[:2] if fallback else headers[:1]


def select_display_columns(headers: list[str]) -> list[str]:
    preferred_keywords = (
        "role",
        "designation",
        "department",
        "section",
        "phone",
        "mobile",
        "ext",
        "resi",
        "residence",
        "email",
        "location",
        "office",
    )
    normalized = {header: normalize_header(header) for header in headers}
    name_columns = set(infer_name_columns(headers))
    selected: list[str] = []
    for keyword in preferred_keywords:
        for header in headers:
            if header in selected or header in name_columns:
                continue
            if keyword in normalized[header]:
                selected.append(header)
                break
    if not selected:
        return [h for h in headers[:4] if h not in name_columns]
    return selected[:6]


def build_search_blob(record: dict[str, str]) -> str:
    values = [value.lower() for key, value in record.items() if key not in {"_score"} and value]
    return " | ".join(values)


def score_record(record: dict[str, str], query_tokens: list[str], name_columns: list[str]) -> int:
    if not query_tokens:
        return 0

    score = 0
    name_text = " ".join(record.get(column, "").lower() for column in name_columns)
    blob = record.get("_search_blob", "")
    for token in query_tokens:
        if token in name_text:
            score += 10
        if token in blob:
            score += 3
        if any(value.lower().startswith(token) for value in record.values() if isinstance(value, str)):
            score += 2
    return score


@lru_cache(maxsize=1)
def _load_directory_cached(workbook_path: str, workbook_mtime: int) -> tuple[list[dict[str, str]], DirectoryMetadata]:
    workbook = load_workbook(filename=workbook_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    header_index, headers = choose_header_row(rows)
    data_rows = rows[header_index + 1 :]
    display_columns = select_display_columns(headers)
    name_columns = infer_name_columns(headers)

    records: list[dict[str, str]] = []
    for row in data_rows:
        values = [stringify(cell) for cell in row]
        if not any(values):
            continue

        record = {
            headers[position]: values[position] if position < len(values) else ""
            for position in range(len(headers))
        }
        record["_display_name"] = next((record.get(column, "") for column in name_columns if record.get(column, "")), "Unnamed Entry")
        role_col = infer_role_column(headers)
        record["_role"] = record.get(role_col, "").strip() if role_col else ""
        record["_search_blob"] = build_search_blob(record)
        records.append(record)

    workbook.close()

    metadata = DirectoryMetadata(
        workbook_name=Path(workbook_path).name,
        sheet_name=worksheet.title,
        total_records=len(records),
        display_columns=display_columns,
    )
    return records, metadata


def load_directory() -> tuple[list[dict[str, str]], DirectoryMetadata]:
    workbook_path = find_workbook()
    stat = workbook_path.stat()
    return _load_directory_cached(str(workbook_path), int(stat.st_mtime_ns))


def search_records(records: list[dict[str, str]], query: str) -> list[dict[str, str]]:
    if not query:
        return records[:MAX_RESULTS]

    query_tokens = [token.lower() for token in query.split() if token.strip()]
    if not query_tokens:
        return records[:MAX_RESULTS]

    headers = [header for header in records[0].keys() if not header.startswith("_")] if records else []
    name_columns = infer_name_columns(headers)

    scored_matches: list[tuple[int, dict[str, str]]] = []
    for record in records:
        score = score_record(record, query_tokens, name_columns)
        if score > 0:
            scored_matches.append((score, record))

    scored_matches.sort(key=lambda item: (-item[0], item[1].get("_display_name", "").lower()))
    return [record for _, record in scored_matches]


app = create_app()


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.getenv("PORT", "5000")), debug=False)